from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Color
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import datetime
import ISMScraper
import requests

#DEFINE CELL COLORS
DARKRED = 'FFC00000'
RED = 'FFFF0000'
ORANGE = 'FFFFC000'
YELLOW = 'FFFFFF00'
GREEN = 'FF00B050'
BLUE = 'FF0070C0'
LIGHTBLUE = 'FF00B0F0'
PURPLE = 'FF7030A0'
WHITE = '00000000'
OTHERWHITE = 'FFFFFFFF'
black = (0,0,0)
white = (255,255,255)
red = (255,0,0)
green = (0,255,0)
blue = (0,0,255)
grey = (128,128,128)


debug = False
debug_CellEditor = True
debug_changeCellTextColor = False
debug_changeCellBackgroundColor = False
debug_changeRow = False
debug_changeColumn = False
debugCellEditor_changeCellValue = True
debugCellEditor_updateDate = False
debugCellEditor_moveTable = True

Sectors_xlsx = "Sectors.xlsx"
Sectors_Directory = "C:/Users/mattb/Dropbox/Github Repositories/AutoStockTrader/Spreadsheets/PMI_NMI/{}".format( Sectors_xlsx ) 

PMI_IndustriesArray = ["(Machinery)","(Computer & Electronic Products)","(Paper Products)",
                    "(Apparel, Leather & Allied Products)","(Printing & Related Support Activities)",
                    "(Primary Metals)","(Nonmetallic Mineral Products)","(Petroleum & Coal Products)",
                    "(Plastics & Rubber Products)","(Miscellaneous Manufacturing)",
                    "(Food, Beverage & Tobacco Products)","(Furniture & Related Products)",
                    "(Transportation Equipment)","(Chemical Products)","(Fabricated Metal Products)",
                    "(Electrical Equipment, Appliances & Components)","(Textile Mills)","(Wood Products)"]
NMI_IndustriesArray = ["(Retail Trade)","(Utilities)","(Arts, Entertainment Recreation)",
                    "(Other Services)","(Healthcare and Social Assistance)","(Food and Accomodations)",
                    "(Finance and Insurance)","(Real Estate, Renting and Leasing)",
                    "(Transport and Warehouse)","(Mining)","(Wholesale)","(Public Admin)",
                    "(Professional, Science and Technology Services)","(Information)","(Education)",
                    "(Management)","(Construction)","(Agriculture, Forest, Fishing and Hunting)"]
PMI_SectorsArray = ["NEW ORDERS","PRODUCTION","EMPLOYMENT","SUPPLIER DELIVERIES","INVENTORIES",
                    "CUSTOMER INVENTORIES","PRICES","BACKLOG OF ORDERS","EXPORTS","IMPORTS"]
NMI_SectorsArray = ["ISM NON-MANUFACTURING","BUSINESS ACTIVITY","NEW ORDERS","EMPLOYMENT","SUPPLIER DELIVERIES","INVENTORIES",
                    "PRICES","BACKLOG OF ORDERS","EXPORTS","IMPORTS"]

MonthArray = [ "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December" ]
IndustryCommentsInitColumn = 68
IndustryCommentsInitRow = 5
IndustryTrendInitColumn = 70
IndustryTrendInitRow = 6

DateRow = 3
DateColumn = 67

def updateURL( newURL ):
    newPage = requests.get( newURL )
    if newPage.status_code == 200:
        print("Request successful!")
        ISM_Page = newPage
        ISM_Soup = ISMScraper.grabSoup( ISM_Page )
        ISM_Type = ISMScraper.grabType( ISM_Soup )    
        if ISM_Type == "PMI":
            thisEditor = PMI_Editor
            thisIndustriesArray = PMI_IndustriesArray
        elif ISM_Type == "NMI":
            thisEditor = NMI_Editor
            thisIndustriesArray = NMI_IndustriesArray
    else:
        print("ERROR: status code: {}".format(newURL.status_code))
        print("Reverting back to previous page...")
        try:
            ISM_Soup = ISMScraper.grabSoup( ISM_Page )
        except NameError:
            print("FATAL ERROR: NO PAGE TO GO BACK TO")
            exit()


    return ISM_Page, ISM_Soup, ISM_Type, thisEditor, thisIndustriesArray

class cellEditor:
    def __init__(self,xlsx,sheetIndex,row=1,column=65):
        self.xlsx = xlsx                                    #e.g. "MultipleTimingSP500PMI_xlsx.xlsx"
        self.workbook = load_workbook(filename = self.xlsx)                    #e.g. load_workbook(filename = "MultipleTimingSP500PMI_xlsx.xlsx")
        if debug_CellEditor == True:
            print("self.workbook: {}".format(self.workbook))
        self.sheet = self.workbook.worksheets[sheetIndex]                                  #e.g. MyDistribution
        if debug_CellEditor == True:
            print("self.sheet: {}".format(self.sheet))
        self.row = row                                      #e.g. 7
        self.column = column                                #e.g. A
        self.coordinates = "{}{}".format(chr(self.column),self.row)   #e.g. 'A7'
        self.cell = self.sheet["{}".format(self.coordinates)]  #e.g. MyDistribution['A7']
        self.value = self.cell.value                        #e.g. "this is the text contained in the current cell"

    #Style change functions
    def changeCellTextColor(self,color):
        self.cell.font = Font(color=color)
        self.saveWorkbook()
        if (debug_changeCellTextColor == True):
            self.printCellFont()
    def changeCellBackgroundColor(self,color):
        self.cell.fill = color
        if (debug_changeCellBackgroundColor == True):
            self.printCellFont()
    def changeCellFont(self, name = "Calibri (Body)", sz = 11, bold = False, italic = False, underline = False, color = black):
        self.cell.font = Font(name=name, sz=sz, bold=bold, italic=italic, underline=underline, color=color)
    #Printing functions
    def printCellCoordinates(self):
        print("Current cell coordinates: {}".format(self.coordinates))
    def printCellValue(self):
        print("Cell {} value is now: {}".format(self.coordinates,self.cell.value))
    def printCellBackgroundColor(self):
        print("Cell {} background color is now: {}".format(self.coordinates,self.cell.fill))
    def printCellFont(self):
        print("Cell {} font is now: {}".format(self.coordinates,self.cell.font))

    #Save xlsx file
    def saveWorkbook(self, thisFileName = 0):
        if thisFileName == 0:
            thisFileName = self.xlsx
        self.workbook.save(filename = "{}".format(self.xlsx))
        print("{} Saved as {}!".format( self.xlsx, thisFileName ))

    #Coordinate change functions
    def cellCoordAdjust(self):
        self.coordinates = "{}{}".format(chr(self.column),self.row)
        self.cell = self.sheet["{}".format(self.coordinates)]
        self.value = self.cell.value
    def changeRow(self,value):
        self.row = value
        self.cellCoordAdjust()
        if (debug_changeRow == True):
            self.printCellCoordinates()
    def changeColumn(self,value):
        self.column = value
        self.cellCoordAdjust()
        if (debug_changeColumn == True):
            self.printCellCoordinates()

    #Value change functions
    def changeCellValue(self,text):
        self.cell.value = text
        if (debugCellEditor_changeCellValue == True):
            self.printCellValue()
    
    def updateDate( self ):
        date = ISMScraper.grabDate( ISM_Soup )
        if debugCellEditor_updateDate == True:
            print("ISMScraper.grabDate( ISM_Soup ): {}".format(date))
        oldrow, oldcolumn = self.row, self.column
        self.row, self.column = DateRow, DateColumn
        self.cellCoordAdjust()
        self.changeCellValue( date )
        self.row, self.column = oldrow, oldcolumn
        self.saveWorkbook()


    def updateComments( self, IndustriesArray, IndustryCommentsCurrentColumn = IndustryCommentsInitColumn, IndustryCommentsCurrentRow = IndustryCommentsInitRow ):
        print("Updating {} comments...".format( self.sheet ))
        self.column = IndustryCommentsCurrentColumn
        self.row = IndustryCommentsCurrentRow
        ISM_CommentList = ISMScraper.grabISMcomments( ISM_Soup )
        for noteIndex, note in enumerate(ISM_CommentList):
            for industryIndex, industry in enumerate(IndustriesArray):
                if debug == True:
                    print("Searching for {} in note {}!".format( industry.strip("()"), noteIndex))
                if note.text.find(industry) != -1:
                    if debug == True:
                        print("note.text.find(industry): {}".format(note.text.find(industry)))
                        print("Found {} in note {}!".format( industry.strip("()"), noteIndex))

                    IndustryCommentsCurrentRow += (2*industryIndex)
                    self.row = IndustryCommentsCurrentRow
                    self.cellCoordAdjust()
                    self.changeCellValue(ISM_CommentList[noteIndex].text)

                    IndustryCommentsCurrentRow = IndustryCommentsInitRow
                    self.row = IndustryCommentsInitRow
                    self.cellCoordAdjust()
                    break

        print("Comments updated successfully! :) Saving...")
        self.saveWorkbook()
        self.saveWorkbook("Backup.xlsx")
        return 0

    def updateRankings( self, IndustriesArray ):
        thisInitColumn = 70
        IndustryRankingsInitRow = 5
        print("Updating PMI rankings...")
        ISM_RankingListArray = [None] * 11
        if ISM_Type == "PMI":
            thisSectorsArray = PMI_SectorsArray
            thisIndustryArray = PMI_IndustriesArray
        elif ISM_Type == "NMI":
            thisSectorsArray = NMI_SectorsArray
            thisIndustryArray = NMI_IndustriesArray
        rankDictionaryArray = ISMScraper.grabISMrankings( ISM_Soup )

        self.column = thisInitColumn
        for rankDictIndex, rankDictionary in enumerate(rankDictionaryArray):
            for currentIndustry, rank in rankDictionary.items():
                self.row = IndustryRankingsInitRow
                # Searching for correct row to put rank into
                for industryIndex, industry in enumerate( thisIndustryArray ):
                    if industry.find(currentIndustry) != -1:
                        self.row += 2*industryIndex
                        print("Found {}!".format(currentIndustry))
                        self.cellCoordAdjust()
                        self.changeCellValue( rank )
                        break
                    
                
            self.column += 1



        ArrayIndex = 0
        print("len(ISM_RankingListArray): {}".format( len (ISM_RankingListArray) ) )
        
        print("Rankings updated successfully! :) Saving...")
        self.saveWorkbook()
        self.saveWorkbook("Backup.xlsx")
        return 0

    #Takes: start cell, end cell, and sheet you want to copy from.
    def copyRange(self, startCol, startRow, endCol, endRow):
        copiedData = []
        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(self.sheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the copiedData
            copiedData.append(rowSelected)
    
        return copiedData

    #Paste data from copyRange into template sheet
    def pasteRange(self, startCol, startRow, endCol, endRow, copiedData):
        sheetReceiving = self.sheet
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1

    def moveTable(self):
        tableBorderLength = 40
        tableBorderWidth = 16
        date = self.sheet["C3"].value
        year = int(date[-4:])
        month = date[:-4]
        if debugCellEditor_moveTable == True:
            print("int(date[-4:]): {}".format(year))
            print("date[:4]: {}".format(month))
        yearDifference = 2025-year
        for monthIndex, monthInstance in enumerate( MonthArray ):
            print("monthInstance: {}".format(monthInstance))
            if monthInstance.find(monthInstance) != -1:
                monthDifference = monthIndex
                break
        tableLengthOffset = tableBorderLength * monthDifference
        tableWidthOffset = tableBorderLength * yearDifference
        copiedData = self.copyRange( 3, 3, 16, 40 )
        self.pasteRange( 3+tableWidthOffset, 3+tableLengthOffset, 16+tableWidthOffset, 40+tableLengthOffset, copiedData )
        if debugCellEditor_moveTable == True:
            print("Copied table {} cells to the right and {} cells down.".format( tableWidthOffset, tableLengthOffset ))

def updateSheetNames( xlsx ):
    print("\nGiving sheet variables to each sheet...")
    if debug == True:
        print("Sheet names for {}:".format( xlsx ))
        for sheet_i, sheet in enumerate(xlsx.worksheets):
            print("Sheet {}: {}".format( sheet_i, sheet ))
    print("Done giving sheet variables to the sheet!")
    return 0


PMI_Editor = cellEditor( Sectors_xlsx, 1 )
NMI_Editor = cellEditor( Sectors_xlsx, 2 )

command = ""
while command != 'q':
    if command == "":
        command = 'nu'
    else:
        print("What would you like to do? ('q'=quit, 'nu'=New URL, 'ua'=Update All, ")
        command = input("'ud'=Update Date, 'uc'=Update Comments, 'ur'=Update Rankings)")
    if command == 'nu':
        newURL = input("Enter new URL here: ")
        ISM_Page, ISM_Soup, ISM_Type, thisEditor, thisIndustriesArray = updateURL( newURL )
    elif command == 'uc':
        thisEditor.updateComments( thisIndustriesArray )
    elif command == 'ur':
        thisEditor.updateRankings( thisIndustriesArray )
    elif command == 'ud':
        thisEditor.updateDate()
    elif command == 'ua':
        thisEditor.updateDate()
        thisEditor.updateComments( thisIndustriesArray )
        thisEditor.updateRankings( thisIndustriesArray )
        # thisEditor.moveTable()
    else:
        print("Invalid command!")
    # elif command == 'md':
        # thisEditor.moveTable()
print("Exiting")